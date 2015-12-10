# -*- coding: utf-8 -*-
from os.path import exists

from hamcrest import assert_that, has_entries, contains_inanyorder as contains
from openpyxl import load_workbook

from tgit import tagging, fs
from tgit.local_storage.local_project import TRACKS_FOLDER_NAME
from tgit.metadata import Image


class AlbumWorkspace(object):
    def __init__(self, local_path):
        self._local_path = local_path

    @property
    def root_path(self):
        return self._local_path.strpath

    def file(self, *paths):
        return self._local_path.join(*paths).strpath

    def contains_track(self, album, filename, front_cover=None, **tags):
        if not exists(self.file(album, TRACKS_FOLDER_NAME, filename)):
            raise AssertionError("Track file '{}' not found in workspace".format(filename))

        track = tagging.load_track(self.file(album, TRACKS_FOLDER_NAME, filename))
        images = []
        # todo use builders and metadata
        if front_cover:
            image, desc = front_cover
            mime = fs.guess_mime_type(image)
            images.append(Image(mime, fs.read(image), type_=Image.FRONT_COVER, desc=desc))

        assert_that(track.metadata, has_entries(tags), "metadata tags")
        assert_that(track.metadata.images, contains(*images), "attached pictures")

    def delete(self):
        self._local_path.remove(ignore_errors=True)

    def contains_soproq_transmission_file(self, filename, track_title, lead_performer, isrc, duration):
        if not exists(self.file(filename)):
            raise AssertionError("SOPROQ file '{}' not found in workspace".format(filename))

        wb = load_workbook(self.file(filename))
        ws = wb.active

        assert_that(ws["M13"].value, track_title, "track title")
        assert_that(ws["N13"].value, lead_performer, "lead performer")
        assert_that(ws["P13"].value, isrc, "isrc")
        assert_that(ws["Q13"].value, duration, "track duration")
