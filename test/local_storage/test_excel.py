from hamcrest import assert_that, equal_to
from openpyxl import load_workbook
import pytest

from test.util import resources


@pytest.yield_fixture
def excel_file(tmpdir):
    folder = tmpdir.mkdir("excel")

    def path(filename):
        return folder.join(filename).strpath

    yield path
    folder.remove()


def test_edit_excel_file(excel_file):
    wb = load_workbook(resources.path("soproq.xlsx"))
    ws = wb.active
    ws["A13"] = "Album title"

    save_to = excel_file("soproq.xlsx")
    wb.save(save_to)

    wb = load_workbook(save_to)
    ws = wb.active
    assert_that(ws["A13"].value, equal_to("Album title"), "the album title")
